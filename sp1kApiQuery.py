from wikiInfo import sp1000Wiki
import alpaca_trade_api as tradeapi
import openpyxl
from myPandas import stockDataFrames
from config import *
import time
import json


def write1000TickersExcel():
    wb = openpyxl.Workbook() #if file is being created first time
    sheet = wb.active

    myTickers = sp1000Wiki.getWiki1000()

    for i, v in enumerate(myTickers):
        cell = 'A' + str((i+2))

        sheet[cell] = v

    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def update1000TickersExcel():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active

    myTickers = sp1000Wiki.getWiki1000()

    for i, v in enumerate(myTickers):
        cell = 'A' + str((i+2))

        sheet[cell] = v

    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def update5YearForAll():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active
    api = getApi()
    myTickers = sp1000Wiki.getWiki1000()
    params = getParams()

    for i,v in enumerate(myTickers):
        try:
            myURL = 'https://api.polygon.io/v1/open-close/{}/{}'.format(v, stockDataFrames.get5Y())
            resp = api.polygon._session.request('GET', myURL, params=params)

            if (i+1) % 100 == 0:
                time.sleep(25)

            jjson = resp.json()
            #print(jjson)
            try:
                rowB = 'B' + str(i+2)
                print(jjson['symbol'])
                print(jjson['low'])
                sheet[rowB] = jjson['low']
            except:
                rowB = 'B' + str(i+2)
                print("Error occured at {}".format(rowB))
                sheet[rowB] = "N/A"
        except:
            sheet[rowB] = "N/A"
            print("Error occured at {}".format(rowB))
            print("Entered \"N\\A\" at {}".format(rowB))
    
    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def update1YearForAll():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active
    api = getApi()
    myTickers = sp1000Wiki.getWiki1000()
    params = getParams()

    for i,v in enumerate(myTickers):
        try:
            myURL = 'https://api.polygon.io/v1/open-close/{}/{}'.format(v, stockDataFrames.get1Y())
            resp = api.polygon._session.request('GET', myURL, params=params)

            if (i+1) % 26 == 0:
                time.sleep(5)

            jjson = resp.json()
            #print(jjson)
            try:
                rowC = 'C' + str(i+2)
                print(jjson['symbol'])
                print(jjson['low'])
                sheet[rowC] = jjson['low']
            except:
                rowC = 'C' + str(i+2)
                print("Error occured at {}".format(rowC))
                sheet[rowC] = "N/A"
        except:
            sheet[rowC] = "N/A"
            print("Error occured at {}".format(rowC))
            print("Entered \"N\\A\" at {}".format(rowC))
    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def updateFeb():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active
    api = getApi()
    myTickers = sp1000Wiki.getWiki1000()
    params = getParams()
    

    for i,v in enumerate(myTickers):
        try:
            myURL = 'https://api.polygon.io/v1/open-close/{}/{}'.format(v, '2020-02-14')
            resp = api.polygon._session.request('GET', myURL, params=params)

            if (i+1) % 26 == 0:
                time.sleep(5)

            jjson = resp.json()
            #print(jjson)
            try:
                rowD = 'D' + str(i+2)
                print(jjson['symbol'])
                print(jjson['low'])
                sheet[rowD] = jjson['low']
            except:
                rowD = 'D' + str(i+2)
                print("Error occured at {}".format(rowD))
                sheet[rowD] = "N/A"
        except:
            rowD = 'D' + str(i+2)
            print("Error occured at {}".format(rowD))
            sheet[rowD] = "N/A"
    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def updateMarch():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active
    api = getApi()
    myTickers = sp1000Wiki.getWiki1000()
    params = getParams()
    

    for i,v in enumerate(myTickers):
        try:
            myURL = 'https://api.polygon.io/v1/open-close/{}/{}'.format(v, '2020-03-20')
            resp = api.polygon._session.request('GET', myURL, params=params)

            if (i+1) % 26 == 0:
                time.sleep(5)

            jjson = resp.json()
            #print(jjson)
            try:
                rowE = 'E' + str(i+2)
                print(jjson['symbol'])
                print(jjson['low'])
                sheet[rowE] = jjson['low']
            except:
                rowE = 'E' + str(i+2)
                print("Error occured at {}".format(rowE))
                sheet[rowE] = "N/A"
        except:
            rowE = 'E' + str(i+2)
            print("Error occured at {}".format(rowE))
            sheet[rowE] = "N/A"

    
    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def update1DayForAll():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active
    api = getApi()
    myTickers = sp1000Wiki.getWiki1000()
    params = getParams()

    for i,v in enumerate(myTickers):
        try:
            myURL = 'https://api.polygon.io/v1/open-close/{}/{}'.format(v, stockDataFrames.get1D())
            resp = api.polygon._session.request('GET', myURL, params=params)

            if (i+1) % 100 == 0:
                time.sleep(25)

            jjson = resp.json()
            #print(jjson)
            try:
                rowF = 'F' + str(i+2)
                print(jjson['symbol'])
                print(jjson['low'])
                sheet[rowF] = jjson['low']
            except:
                rowF = 'F' + str(i+2)
                print("Error occured at {}".format(rowF))
                sheet[rowF] = "N/A"
        except:
            rowF = 'F' + str(i+2)
            print("Error occured at {}".format(rowF))
            sheet[rowF] = "N/A"
    
    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

#After opening:
def openMarket():
    wb = openpyxl.load_workbook('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx') #if file exists
    sheet = wb.active
    api = getApi()
    myTickers = sp1000Wiki.getWiki1000()
    params = getParams()
    
    for i,v in enumerate(myTickers):
        myURL = 'https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/tickers/{}'.format(v)
        try:
            resp = api.polygon._session.request('GET', myURL, params=params)

            if (i+1) % 100 == 0:
                time.sleep(5)

            jjson = resp.json()
            #print(jjson)
            jjson = resp.json()
            print(json.dumps(jjson, indent=4))
            try:
                rowG = 'G' + str(i+2)
                if jjson['ticker']['lastTrade']['p'] == 0:
                    print('Val was 0')
                else:
                    sheet[rowG] = jjson['ticker']['lastTrade']['p']
            except:
                rowG = 'G' + str(i+2)
                print("Error occured at {}".format(rowG))
                sheet[rowG] = "N/A"
        except:
            rowG = 'G' + str(i+2)
            print("Error occured at {}".format(rowG))
            sheet[rowG] = "N/A"
    
    wb.save('C:/Users/aj282/OneDrive/Desktop/pythonStocks/my1000Xl.xlsx')

def getApi():
    api = tradeapi.REST(API_KEY, SECRET_KEY, api_version='v1')
    return api

def getParams():
    api = tradeapi.REST(API_KEY, SECRET_KEY, api_version='v1')
    params={'apiKey' : api.polygon._api_key}
    return params

if __name__ == '__main__':
    write1000TickersExcel()
    update1000TickersExcel()
    update5YearForAll()
    update1YearForAll()
    updateFeb()
    updateMarch()
    update1DayForAll()
    openMarket()
    #closingPrice()
    #afterHours()
    #preMarket()